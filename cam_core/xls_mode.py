from typing import Dict, Any, List, Tuple
import argparse, datetime, os, re
import shlex
from cam_core.debug import debug_dump_params_and_dir

def parse_dir_listing(text: str) -> list[str]:
    """
    Convert a text dump of files (with directory headers and multiple
    columns per line) into a flat list of full path strings.
    """
    results = []
    current_dir = ""

    if not text:
        print("bad text to parse")
        return

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        # Directory header (ends with ':')
        if line.endswith(":"):
            current_dir = line.rstrip(":")
            continue

        # The shlex.split() function correctly handles the quoted sections
        # remove leading"./"
        line = line.lstrip("./")
        #print("Splitting line: "+line)
        result = re.split(r"(?<!\\) ", line)
        for token in result:
            token.replace("\\\\","")
            results.append(f"{token}")
            #print("     "+token)

    return results
def process_xls(inpath: str, planner_func, verbose: bool=True):
    # copy input to tests_out.xlsx
    out_copy = os.path.join(os.path.dirname(inpath), "tests_out.xlsx")
    try:
        import shutil as _sh
        _sh.copyfile(inpath, out_copy)
        xls_path = out_copy
        print(f"[xls] copied input to: {xls_path}")
    except Exception as e:
        print(f"[xls] ERROR copy failed: {e}")
        return 0
    # verbose parameter + dir dump

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        print("[error] XLS mode requires openpyxl. Install with: pip install openpyxl")
        return 1

    wb = openpyxl.load_workbook(xls_path)
    ws = wb.active

    headers = { (cell.value or "").strip().lower(): cell.column for cell in ws[1] if cell.value }
    def col(name, default=None):
        return headers.get(name.lower(), default)

    col_config = col("Config File")
    col_files = col("Directory Listing")
    col_last_date = col("Last Test Date") or (ws.max_column + 1); ws.cell(row=1, column=col_last_date, value="Last Test Date")
    col_last_out = col("Last Test Output") or (ws.max_column + 1); ws.cell(row=1, column=col_last_out, value="Last Output")
    col_last_log = col("Last Test Log") or (ws.max_column + 1); ws.cell(row=1, column=col_last_log, value="Last Log")
    print(f"[xls] cfg col: ", col_config)
    print(f"[xls] files col: ", col_files)
    print(f"[xls] date col: ", col_last_date)
    print(f"[xls] out col: ", col_last_out)
    print(f"[xls] log col: ", col_last_log)

    from .jsonc_loader import load_config_text, normalize_legacy

    for r in range(2, ws.max_row + 1):

        print("==================ROW "+str(r)+"=====================")
        cfg_text = ws.cell(row=r, column=col_config).value if col_config else None
        vfiles = ws.cell(row=r, column=col_files).value if col_files else None
        if 0:
            print("======================CFG==================")
            print(cfg_text)
            print("==============================================")
        if 0:
            print("======================VFILES==================")
            print(vfiles)
            print("==============================================")

        files = parse_dir_listing(vfiles)
        if 0:
            print("======================FILES==================")
            print(files)
            print("==============================================")


        cfg_dict = {}
        if cfg_text:
            try:
                cfg_dict = normalize_legacy(load_config_text(str(cfg_text)))
            except Exception as e:
                print(f"[warn] row {r} config parse error: {e}")
                cfg_dict = {}

        from .planner import plan as core_plan
        params = {p["name"]: p.get("default") for p in cfg_dict.get("parameters",[]) if isinstance(p, dict) and p.get("name")}

        #debug_dump_params_and_dir(".", cfg_dict, files)

        resolved, by_step = core_plan(str("."), cfg_dict, params, files, verbose=False)


        parts = ""
        for out in resolved:
            step = str(out.get("step",""))
            name = out.get("name","")
            files = by_step.get(step, [])

            if not files:
                continue

            parts= parts + name + ":"

            first=True
            for f in files:
                if not first:
                    parts = parts + ","
                #    parts.append(",")
                parts = parts + f
                first=False
            parts =parts + "\n"

        print(parts)
        ws.cell(row=r, column=col_last_date, value=datetime.datetime.now().isoformat(timespec="seconds"))
        ws.cell(row=r, column=col_last_out, value=", " + parts)
        ws.cell(row=r, column=col_last_log, value="ok" if parts else "no matches")

    wb.save(xls_path)
    print(f"[xls] updated: {xls_path}")
    return 0
